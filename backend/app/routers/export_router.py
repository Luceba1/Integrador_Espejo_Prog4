from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Annotated, Iterable

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.auth_dependencies import UowDep, require_roles
from app.models.usuario import Usuario
from app.uow.unit_of_work import SQLModelUnitOfWork

router = APIRouter(prefix="/export", tags=["Exportación Excel"])
AdminDep = Annotated[Usuario, Depends(require_roles("ADMIN"))]


def _fmt(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, Decimal):
        return float(value)
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sí" if value else "No"
    return value


def _precio_por_unidad(ingrediente) -> Decimal:
    unitario = Decimal(getattr(ingrediente, "precio_costo_unitario", 0) or 0)
    if unitario > 0:
        return unitario
    stock = Decimal(ingrediente.stock_cantidad or 0)
    precio_total = Decimal(getattr(ingrediente, "precio_costo_total", 0) or 0)
    if stock <= 0 or precio_total <= 0:
        return Decimal("0.00")
    return precio_total / stock


def _costos_producto(uow: SQLModelUnitOfWork, producto) -> tuple[Decimal, Decimal]:
    costo = Decimal("0.00")
    for config in uow.productos.list_ingrediente_config(producto.id):
        ingrediente = uow.ingredientes.get_by_id_with_unit(config.ingrediente_id)
        if ingrediente:
            costo += _precio_por_unidad(ingrediente) * Decimal(config.cantidad or 0)
    costo = costo.quantize(Decimal("0.01"))
    margen = Decimal(getattr(producto, "margen_ganancia_porcentaje", 0) or 0)
    sugerido = (costo * (Decimal("1") + margen / Decimal("100"))).quantize(Decimal("0.01"))
    return costo, sugerido


def _xlsx_response(filename: str, headers: list[str], rows: Iterable[Iterable[object]]) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(headers)
    for row in rows:
        ws.append([_fmt(value) for value in row])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 45)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    ws.freeze_panes = "A2"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/productos.xlsx")
def exportar_productos(
    uow: UowDep,
    _: AdminDep,
    incluir_eliminados: Annotated[bool, Query()] = True,
):
    productos = uow.productos.list_with_relations(incluir_eliminados=incluir_eliminados, page=1, size=10000)
    rows = []
    for producto in productos:
        categorias = ", ".join(categoria.nombre for categoria in producto.categorias)
        ingredientes = []
        for config in uow.productos.list_ingrediente_config(producto.id):
            ingrediente = uow.ingredientes.get_by_id_with_unit(config.ingrediente_id)
            unidad = uow.unidades_medida.get_by_id(config.unidad_medida_id) if config.unidad_medida_id else None
            if ingrediente:
                ingredientes.append(f"{ingrediente.nombre}: {config.cantidad} {unidad.simbolo if unidad else ''}".strip())
        costo_ingredientes, precio_sugerido = _costos_producto(uow, producto)
        rows.append([
            producto.id,
            producto.nombre,
            producto.descripcion,
            producto.precio_base,
            costo_ingredientes,
            producto.margen_ganancia_porcentaje,
            precio_sugerido,
            producto.disponible,
            producto.activo,
            "Eliminado" if producto.deleted_at else "Activo",
            producto.unidad_venta.simbolo if producto.unidad_venta else "",
            categorias,
            "; ".join(ingredientes),
            len(producto.imagenes_url or []),
            producto.created_at,
            producto.updated_at,
            producto.deleted_at,
        ])
    return _xlsx_response(
        "productos.xlsx",
        ["ID", "Nombre", "Descripción", "Precio", "Costo ingredientes", "Margen %", "Precio sugerido", "Disponible", "Activo", "Estado", "Unidad venta", "Categorías", "Receta/ingredientes", "Imágenes", "Creado", "Actualizado", "Eliminado"],
        rows,
    )


@router.get("/categorias.xlsx")
def exportar_categorias(
    uow: UowDep,
    _: AdminDep,
    incluir_eliminadas: Annotated[bool, Query()] = True,
):
    categorias = uow.categorias.list_all_without_pagination(incluir_eliminadas=incluir_eliminadas)
    by_id = {categoria.id: categoria for categoria in categorias}
    rows = []
    for categoria in categorias:
        padre = by_id.get(categoria.parent_id) if categoria.parent_id else None
        rows.append([
            categoria.id,
            categoria.nombre,
            categoria.descripcion,
            padre.nombre if padre else "Raíz",
            categoria.parent_id,
            categoria.activo,
            "Eliminada" if categoria.deleted_at else "Activa",
            categoria.imagen_url,
            categoria.created_at,
            categoria.updated_at,
            categoria.deleted_at,
        ])
    return _xlsx_response(
        "categorias.xlsx",
        ["ID", "Nombre", "Descripción", "Padre", "Padre ID", "Activa", "Estado", "Imagen", "Creada", "Actualizada", "Eliminada"],
        rows,
    )


@router.get("/ingredientes.xlsx")
def exportar_ingredientes(
    uow: UowDep,
    _: AdminDep,
    incluir_eliminados: Annotated[bool, Query()] = True,
):
    ingredientes = uow.ingredientes.list_all_with_units(incluir_eliminados=incluir_eliminados)
    rows = [
        [
            item.id,
            item.nombre,
            item.descripcion,
            item.es_alergeno,
            item.stock_cantidad,
            item.precio_costo_total,
            _precio_por_unidad(item).quantize(Decimal("0.01")),
            item.unidad_medida.simbolo if item.unidad_medida else "",
            item.unidad_medida.nombre if item.unidad_medida else "",
            item.activo,
            "Eliminado" if item.deleted_at else "Activo",
            item.created_at,
            item.updated_at,
            item.deleted_at,
        ]
        for item in ingredientes
    ]
    return _xlsx_response(
        "ingredientes.xlsx",
        ["ID", "Nombre", "Descripción", "Alérgeno", "Stock", "Precio total", "Precio x unidad", "Unidad", "Unidad nombre", "Activo", "Estado", "Creado", "Actualizado", "Eliminado"],
        rows,
    )


@router.get("/unidades-medida.xlsx")
def exportar_unidades_medida(
    uow: UowDep,
    _: AdminDep,
    incluir_eliminadas: Annotated[bool, Query()] = True,
):
    unidades = uow.unidades_medida.list_all_ordered(incluir_eliminadas=incluir_eliminadas)
    rows = [
        [item.id, item.nombre, item.simbolo, item.tipo, item.activo, "Eliminada" if item.deleted_at else "Activa", item.created_at, item.updated_at, item.deleted_at]
        for item in unidades
    ]
    return _xlsx_response(
        "unidades_medida.xlsx",
        ["ID", "Nombre", "Símbolo", "Tipo", "Activa", "Estado", "Creada", "Actualizada", "Eliminada"],
        rows,
    )


@router.get("/usuarios.xlsx")
def exportar_usuarios(
    uow: UowDep,
    _: AdminDep,
    incluir_eliminados: Annotated[bool, Query()] = True,
):
    usuarios = uow.usuarios.list_paginated_admin(incluir_eliminados=incluir_eliminados, page=1, size=10000)
    rows = [
        [
            usuario.id,
            usuario.email,
            usuario.nombre,
            usuario.apellido,
            usuario.celular,
            usuario.activo,
            "Eliminado" if usuario.deleted_at else "Activo",
            ", ".join(rol.codigo for rol in usuario.roles),
            usuario.created_at,
            usuario.updated_at,
            usuario.deleted_at,
        ]
        for usuario in usuarios
    ]
    return _xlsx_response(
        "usuarios.xlsx",
        ["ID", "Email", "Nombre", "Apellido", "Celular", "Activo", "Estado", "Roles", "Creado", "Actualizado", "Eliminado"],
        rows,
    )
